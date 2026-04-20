"""
审核报告生成模块
使用 openpyxl 生成标记版 (.xlsx) 和详情版 (.xlsx) 审核报告，
以及批量打包 (.zip) 功能。

优化：标记版报告能还原 PDF 的多列表格排版，而非堆成一列。
"""

import io
import logging
import re
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ============================================================
# 颜色 / 样式常量
# ============================================================
_FILL_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
_FILL_BLUE = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FILL_HEADER_LIGHT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_FILL_DETAIL_RED = PatternFill(start_color="FFF1F0", end_color="FFF1F0", fill_type="solid")
_FILL_DETAIL_YELLOW = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
_FILL_DETAIL_BLUE = PatternFill(start_color="E6F4FF", end_color="E6F4FF", fill_type="solid")

_FONT_TITLE = Font(name="Microsoft YaHei", size=16, bold=True, color="1A1A1A")
_FONT_HEADER = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
_FONT_HEADER_DARK = Font(name="Microsoft YaHei", size=11, bold=True, color="1A1A1A")
_FONT_NORMAL = Font(name="Microsoft YaHei", size=10, color="333333")
_FONT_BOLD = Font(name="Microsoft YaHei", size=10, bold=True, color="333333")
_FONT_RED = Font(name="Microsoft YaHei", size=10, bold=True, color="CF1322")
_FONT_YELLOW = Font(name="Microsoft YaHei", size=10, bold=True, color="D48806")
_FONT_BLUE = Font(name="Microsoft YaHei", size=10, bold=True, color="1677FF")
_FONT_GRAY = Font(name="Microsoft YaHei", size=9, color="888888")
_FONT_GREEN = Font(name="Microsoft YaHei", size=11, bold=True, color="008000")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

_LEVEL_FILL_MAP = {
    "RED": _FILL_RED,
    "YELLOW": _FILL_YELLOW,
    "BLUE": _FILL_BLUE,
}

_LEVEL_DETAIL_FILL_MAP = {
    "RED": _FILL_DETAIL_RED,
    "YELLOW": _FILL_DETAIL_YELLOW,
    "BLUE": _FILL_DETAIL_BLUE,
}

_LEVEL_FONT_MAP = {
    "RED": _FONT_RED,
    "YELLOW": _FONT_YELLOW,
    "BLUE": _FONT_BLUE,
}

_LEVEL_LABEL_MAP = {
    "RED": "高风险",
    "YELLOW": "需注意",
    "BLUE": "格式提醒",
}

_LEVEL_SORT_ORDER = {"RED": 0, "YELLOW": 1, "BLUE": 2}

# 标记版最大列数
_MAX_COLS = 15


# ============================================================
# 辅助函数
# ============================================================
def _today_str() -> str:
    return datetime.now().strftime("%Y%m%d")


def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def _extract_contract_no(audit_issues: Optional[List[Dict]] = None,
                         fallback: str = "未知") -> str:
    if audit_issues:
        for issue in audit_issues:
            fname = issue.get("field_name", "")
            if "合同" in fname or "invoice" in fname.lower() or "contract" in fname.lower():
                val = issue.get("source_value", "")
                if val:
                    return _sanitize_filename(str(val))
    return fallback


def _auto_adjust_column_widths(ws, min_width: int = 8, max_width: int = 50) -> None:
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                text = str(cell.value)
                for line in text.split('\n'):
                    line_len = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in line)
                    max_len = max(max_len, line_len + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def _apply_border_to_range(ws, min_row, max_row, min_col, max_col) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _THIN_BORDER


def _find_best_matching_line(issue: Dict, lines: List[str]) -> int:
    """为一个 issue 找到原文中最匹配的行号。"""
    your_value = str(issue.get("your_value", "")).strip()
    field_name = str(issue.get("field_name", "")).strip()
    location = str(issue.get("field_location", "")).strip()

    if your_value and len(your_value) >= 2:
        for idx, line in enumerate(lines):
            if your_value in line:
                return idx

    if field_name:
        keywords = field_name.replace("（", " ").replace("）", " ").split()
        for idx, line in enumerate(lines):
            if any(kw in line for kw in keywords if len(kw) >= 2):
                return idx

    if location:
        for idx, line in enumerate(lines):
            if location in line:
                return idx

    return max(0, len(lines) - 1)


def _parse_table_line(line: str) -> Optional[List[str]]:
    """尝试将 '| col1 | col2 | col3 |' 格式的行解析为列表。"""
    line = line.strip()
    if line.startswith("|") and line.endswith("|"):
        parts = line[1:-1].split("|")
        return [p.strip() for p in parts]
    return None


# ============================================================
# 标记版报告生成 (Excel) — 优化版：保留表格多列结构
# ============================================================
def generate_marked_report(
    original_text: str,
    audit_issues: List[Dict[str, Any]],
    doc_type: str,
    contract_no: str,
    structured_data: Optional[Dict] = None,
) -> Tuple[io.BytesIO, str]:
    """生成标记版审核报告 (.xlsx)。

    优化：将原始 PDF 的表格多列结构还原到 Excel，而非全部堆成单列。

    Args:
        original_text: 单据的原始解析文本。
        audit_issues: 审核发现的问题列表。
        doc_type: 单据类型名称。
        contract_no: 合同号/发票号。
        structured_data: PDF 结构化数据（可选，来自 file_parser.parse_pdf_structured）。

    Returns:
        (BytesIO 对象, 文件名) 的元组。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "审核标记版"

    date_str = _today_str()
    safe_contract = _sanitize_filename(contract_no)
    safe_doc_type = _sanitize_filename(doc_type)
    filename = f"审核标记_{safe_doc_type}_{safe_contract}_{date_str}.xlsx"

    # --- 标题区 ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"审核标记报告 - {doc_type}"
    title_cell.font = _FONT_TITLE
    title_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    info_cell = ws["A2"]
    info_cell.value = f"合同号：{contract_no}    审核日期：{date_str}"
    info_cell.font = _FONT_GRAY
    info_cell.alignment = _ALIGN_CENTER

    ws.merge_cells("A3:H3")
    legend_cell = ws["A3"]
    legend_cell.value = (
        "标记图例：  红色底色 = 高风险(RED)    黄色底色 = 需注意(YELLOW)    "
        "蓝色底色 = 格式提醒(BLUE)    （有问题的单元格附带批注，请鼠标悬停查看）"
    )
    legend_cell.font = _FONT_GRAY
    legend_cell.alignment = _ALIGN_LEFT

    current_row = 5

    # --- 解析原始文本为行 ---
    lines = original_text.split("\n") if original_text else ["（无原始文本）"]

    # 为每个 issue 找到最匹配的行
    line_issues: Dict[int, List[Dict]] = {}
    for issue in audit_issues:
        best_line = _find_best_matching_line(issue, lines)
        if best_line not in line_issues:
            line_issues[best_line] = []
        line_issues[best_line].append(issue)

    # --- 逐行写入 ---
    for line_idx, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            current_row += 1
            continue

        # 尝试解析为表格行
        table_cells = _parse_table_line(stripped)

        if table_cells:
            # 写入为表格行，每个单元格一列（保留多列结构）
            num_cols = min(len(table_cells), _MAX_COLS)
            for col_idx, cell_val in enumerate(table_cells[:num_cols], start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=cell_val)
                cell.font = _FONT_NORMAL
                cell.alignment = _ALIGN_LEFT
                cell.border = _THIN_BORDER
        else:
            # 普通文本行
            if stripped.startswith("=") and ("页" in stripped or "工作表" in stripped):
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=8)
                cell = ws.cell(row=current_row, column=1, value=stripped)
                cell.font = _FONT_BOLD
                cell.fill = _FILL_HEADER_LIGHT
                cell.alignment = _ALIGN_CENTER
            elif stripped.startswith("[表格") or stripped.startswith("--- "):
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=8)
                cell = ws.cell(row=current_row, column=1, value=stripped)
                cell.font = _FONT_BOLD
                cell.alignment = _ALIGN_LEFT
            else:
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=8)
                cell = ws.cell(row=current_row, column=1, value=stripped)
                cell.font = _FONT_NORMAL
                cell.alignment = _ALIGN_LEFT

        # --- 添加问题标记（背景色 + 批注）---
        if line_idx in line_issues:
            issues_for_line = line_issues[line_idx]

            highest_level = "BLUE"
            for iss in issues_for_line:
                lv = iss.get("level", "YELLOW")
                if lv == "RED":
                    highest_level = "RED"
                    break
                elif lv == "YELLOW" and highest_level != "RED":
                    highest_level = "YELLOW"

            fill = _LEVEL_FILL_MAP.get(highest_level, _FILL_YELLOW)

            comment_parts = []
            for iss in issues_for_line:
                iss_id = iss.get("id", "?")
                iss_level = iss.get("level", "YELLOW")
                iss_label = _LEVEL_LABEL_MAP.get(iss_level, "需注意")
                field_name = iss.get("field_name", "")
                suggestion = iss.get("suggestion", "")
                your_val = iss.get("your_value", "")
                src_val = iss.get("source_value", "")

                part = f"[{iss_id}] {iss_label} - {field_name}"
                if your_val or src_val:
                    part += f"\n  单据值: {your_val}\n  PO原值: {src_val}"
                if suggestion:
                    part += f"\n  建议: {suggestion}"
                comment_parts.append(part)

            comment_text = "\n\n".join(comment_parts)
            comment = Comment(comment_text, "审核系统")
            comment.width = 400
            comment.height = max(100, len(comment_parts) * 80)

            if table_cells:
                num_cols = min(len(table_cells), _MAX_COLS)
                for col_idx in range(1, num_cols + 1):
                    ws.cell(row=current_row, column=col_idx).fill = fill
                ws.cell(row=current_row, column=1).comment = comment
            else:
                ws.cell(row=current_row, column=1).fill = fill
                ws.cell(row=current_row, column=1).comment = comment

        current_row += 1

    # --- 页脚 ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=8)
    footer = ws.cell(row=current_row, column=1)
    footer.value = "本报告由「外贸跟单工单智能审核系统」自动生成 | AI审核结果仅供参考，请务必人工复核"
    footer.font = _FONT_GRAY
    footer.alignment = _ALIGN_CENTER

    _auto_adjust_column_widths(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer, filename


# ============================================================
# 详情版报告生成 (Excel)
# ============================================================
def generate_detail_report(
    audit_issues: List[Dict[str, Any]],
    doc_type: str,
    contract_no: str,
) -> Tuple[io.BytesIO, str]:
    """生成审核详情报告 (.xlsx)。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "审核详情"

    date_str = _today_str()
    safe_contract = _sanitize_filename(contract_no)
    safe_doc_type = _sanitize_filename(doc_type)
    filename = f"审核详情_{safe_doc_type}_{safe_contract}_{date_str}.xlsx"

    red_count = sum(1 for i in audit_issues if i.get("level") == "RED")
    yellow_count = sum(1 for i in audit_issues if i.get("level") == "YELLOW")
    blue_count = sum(1 for i in audit_issues if i.get("level") == "BLUE")
    total_count = len(audit_issues)

    # 标题区
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "审核详情报告"
    title_cell.font = _FONT_TITLE
    title_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"单据类型：{doc_type}    合同号：{contract_no}    审核日期：{date_str}"
    ws["A2"].font = _FONT_GRAY
    ws["A2"].alignment = _ALIGN_CENTER

    # 总览区
    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="一、审核总览").font = Font(
        name="Microsoft YaHei", size=13, bold=True, color="1A1A1A"
    )
    row += 1

    summary_headers = ["级别", "数量", "说明"]
    summary_data = [
        ("RED - 高风险", str(red_count), "可能直接导致经济损失或清关失败"),
        ("YELLOW - 需注意", str(yellow_count), "信息存在差异但可合理解释，需人工确认"),
        ("BLUE - 格式提醒", str(blue_count), "纯格式或排版建议，不影响实际业务"),
    ]
    summary_fills = [_FILL_DETAIL_RED, _FILL_DETAIL_YELLOW, _FILL_DETAIL_BLUE]
    summary_fonts = [_FONT_RED, _FONT_YELLOW, _FONT_BLUE]

    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
    row += 1

    for i, (level_name, count, desc) in enumerate(summary_data):
        ws.cell(row=row, column=1, value=level_name).font = summary_fonts[i]
        ws.cell(row=row, column=1).fill = summary_fills[i]
        ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=1).border = _THIN_BORDER

        ws.cell(row=row, column=2, value=count).font = _FONT_BOLD
        ws.cell(row=row, column=2).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=2).fill = summary_fills[i]
        ws.cell(row=row, column=2).border = _THIN_BORDER

        ws.cell(row=row, column=3, value=desc).font = _FONT_NORMAL
        ws.cell(row=row, column=3).fill = summary_fills[i]
        ws.cell(row=row, column=3).alignment = _ALIGN_LEFT
        ws.cell(row=row, column=3).border = _THIN_BORDER
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"共发现 {total_count} 处标记").font = _FONT_BOLD
    ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
    row += 2

    # 逐项详情
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="二、逐项审核详情").font = Font(
        name="Microsoft YaHei", size=13, bold=True, color="1A1A1A"
    )
    row += 1

    if not audit_issues:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value="未发现任何问题，该单据审核通过。").font = _FONT_GREEN
        ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
    else:
        detail_headers = [
            "问题编号", "风险等级", "所在位置", "字段名称",
            "PO原始值", "实际值", "问题说明", "修改建议"
        ]
        for col_idx, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border = _THIN_BORDER
        row += 1

        sorted_issues = sorted(
            audit_issues,
            key=lambda x: _LEVEL_SORT_ORDER.get(x.get("level", "YELLOW"), 1)
        )

        for issue in sorted_issues:
            level = issue.get("level", "YELLOW")
            issue_id = issue.get("id", "?")
            field_location = issue.get("field_location", "")
            field_name = issue.get("field_name", "")
            source_value = issue.get("source_value", "")
            your_value = issue.get("your_value", "")
            source = issue.get("source", "")
            suggestion = issue.get("suggestion", "")

            level_label = _LEVEL_LABEL_MAP.get(level, "需注意")
            level_fill = _LEVEL_DETAIL_FILL_MAP.get(level, _FILL_DETAIL_YELLOW)
            level_font = _LEVEL_FONT_MAP.get(level, _FONT_YELLOW)

            description = f"数据来源: {source}" if source else ""

            row_data = [
                issue_id,
                f"{level} ({level_label})",
                field_location,
                field_name,
                str(source_value),
                str(your_value),
                description,
                suggestion,
            ]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = _FONT_NORMAL
                cell.alignment = _ALIGN_LEFT_TOP
                cell.border = _THIN_BORDER
                cell.fill = level_fill

            ws.cell(row=row, column=2).font = level_font
            ws.cell(row=row, column=1).font = _FONT_BOLD
            ws.cell(row=row, column=1).alignment = _ALIGN_CENTER

            row += 1

    # 页脚
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1,
            value="本报告由「外贸跟单工单智能审核系统」自动生成").font = _FONT_GRAY
    ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1,
            value="AI审核结果仅供参考，请务必进行人工复核").font = Font(
        name="Microsoft YaHei", size=9, color="CF1322"
    )
    ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=f"生成时间：{date_str}").font = _FONT_GRAY
    ws.cell(row=row, column=1).alignment = _ALIGN_CENTER

    _auto_adjust_column_widths(ws, min_width=10, max_width=40)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 35

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer, filename


# ============================================================
# ZIP 打包
# ============================================================
def generate_zip(
    all_reports: List[Tuple[str, io.BytesIO]],
    contract_no: str = "未知",
) -> Tuple[io.BytesIO, str]:
    """将多份报告打包为 ZIP 文件。"""
    date_str = _today_str()
    safe_contract = _sanitize_filename(contract_no)
    zip_filename = f"审核报告_{safe_contract}_{date_str}.zip"

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in all_reports:
            fbytes.seek(0)
            zf.writestr(fname, fbytes.read())

    zip_buffer.seek(0)
    return zip_buffer, zip_filename
